"""
tests/test_export.py
Testes de integracao leve para core/export.py.

Verifica que:
- build_grouped_summary agrupa e soma corretamente.
- to_excel_bytes_grouped gera bytes validos de .xlsx com as duas abas.
- Linha TOTAL esta presente na aba de resumo.
- Colunas numericas sao somadas corretamente no agrupamento.
"""
import io

import openpyxl
import pandas as pd
import pytest

from core.export import build_grouped_summary, to_excel_bytes_grouped


@pytest.fixture
def df_exemplo():
    return pd.DataFrame({
        "Data": pd.to_datetime(["2024-01-01", "2024-01-01", "2024-01-02", "2024-01-02"]),
        "Oficina": ["Alpha", "Alpha", "Beta", "Alpha"],
        "Pecas": [100, 200, 300, 50],
        "Minutos": [60.0, 120.0, 180.0, 30.0],
    })


# --------------------------------------------------------------------------- #
# build_grouped_summary
# --------------------------------------------------------------------------- #
class TestBuildGroupedSummary:
    def test_agrupa_mesma_data_oficina(self, df_exemplo):
        result = build_grouped_summary(df_exemplo, "Data", "Oficina", ["Pecas"])
        row = result[
            (result["Oficina"] == "Alpha") & (result["Data"] == pd.Timestamp("2024-01-01"))
        ]
        assert row["Pecas"].iloc[0] == 300  # 100 + 200

    def test_numero_de_linhas(self, df_exemplo):
        result = build_grouped_summary(df_exemplo, "Data", "Oficina", ["Pecas"])
        assert len(result) == 3  # Alpha/01-01, Alpha/01-02, Beta/01-02

    def test_soma_multiplas_colunas(self, df_exemplo):
        result = build_grouped_summary(df_exemplo, "Data", "Oficina", ["Pecas", "Minutos"])
        row = result[
            (result["Oficina"] == "Alpha") & (result["Data"] == pd.Timestamp("2024-01-01"))
        ]
        assert row["Pecas"].iloc[0] == 300
        assert row["Minutos"].iloc[0] == 180.0

    def test_ordenado_por_data_oficina(self, df_exemplo):
        result = build_grouped_summary(df_exemplo, "Data", "Oficina", ["Pecas"])
        datas = result["Data"].tolist()
        assert datas == sorted(datas)

    def test_df_vazio(self):
        df = pd.DataFrame({
            "Data": pd.Series([], dtype="datetime64[ns]"),
            "Oficina": pd.Series([], dtype=str),
            "Pecas": pd.Series([], dtype=int),
        })
        result = build_grouped_summary(df, "Data", "Oficina", ["Pecas"])
        assert result.empty


# --------------------------------------------------------------------------- #
# to_excel_bytes_grouped
# --------------------------------------------------------------------------- #
class TestToExcelBytesGrouped:
    def test_retorna_bytes(self, df_exemplo):
        data = to_excel_bytes_grouped(df_exemplo, "Data", "Oficina", int_cols=["Pecas"])
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_xlsx_valido(self, df_exemplo):
        data = to_excel_bytes_grouped(df_exemplo, "Data", "Oficina", int_cols=["Pecas"])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert wb is not None

    def test_duas_abas(self, df_exemplo):
        data = to_excel_bytes_grouped(df_exemplo, "Data", "Oficina", int_cols=["Pecas"])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Resumo Diario" in wb.sheetnames
        assert "Detalhado" in wb.sheetnames

    def test_linha_total_presente(self, df_exemplo):
        data = to_excel_bytes_grouped(df_exemplo, "Data", "Oficina", int_cols=["Pecas"])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Resumo Diario"]
        last_row = ws.max_row
        assert ws.cell(row=last_row, column=1).value == "TOTAL"

    def test_aba_detalhada_tem_todos_registros(self, df_exemplo):
        data = to_excel_bytes_grouped(df_exemplo, "Data", "Oficina", int_cols=["Pecas"])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["Detalhado"]
        assert ws.max_row == 5  # 1 cabecalho + 4 linhas de dados

    def test_nomes_de_abas_customizados(self, df_exemplo):
        data = to_excel_bytes_grouped(
            df_exemplo, "Data", "Oficina",
            int_cols=["Pecas"],
            summary_sheet="Resumo",
            detail_sheet="Detalhe",
        )
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Resumo" in wb.sheetnames
        assert "Detalhe" in wb.sheetnames

    def test_com_float_cols(self, df_exemplo):
        data = to_excel_bytes_grouped(
            df_exemplo, "Data", "Oficina",
            int_cols=["Pecas"],
            float_cols=["Minutos"],
        )
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Resumo Diario" in wb.sheetnames
