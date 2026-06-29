"""
core/export.py
Exportacao para Excel (.xlsx) com duas abas: resumo agrupado por Dia+Oficina
(com linha de TOTAL via formula SUM) e dados detalhados registro a registro.

Visual identico ao tema do app: cabecalho verde-cyano, zebra striping, bordas
finas, cabecalho fixo (freeze panes) e auto-filtro.
"""
from __future__ import annotations

import io
from typing import Iterable, Optional

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# Paleta do tema para o Excel
# --------------------------------------------------------------------------- #
_HEADER_FILL = "2EE6C0"
_HEADER_FONT_COLOR = "06151A"
_ZEBRA_FILL = "EAF9F4"
_BORDER_COLOR = "D7E4E1"
_TOTAL_FILL = "CFEFE6"


# --------------------------------------------------------------------------- #
# Helpers internos de estilo openpyxl
# --------------------------------------------------------------------------- #
def _style_worksheet(ws, n_data_rows: int, n_cols: int, total_row: Optional[int] = None) -> None:
    """Cabecalho destacado, zebra striping, bordas, freeze pane e auto-filtro."""
    thin = Side(style="thin", color=_BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    header_font = Font(name="Calibri", bold=True, color=_HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 22

    for row in range(2, n_data_rows + 2):
        is_total = total_row is not None and row == total_row
        is_zebra = (row % 2 == 0) and not is_total
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = center
            cell.border = border
            if is_total:
                cell.fill = PatternFill("solid", fgColor=_TOTAL_FILL)
                cell.font = Font(name="Calibri", bold=True, size=10.5)
            elif is_zebra:
                cell.fill = PatternFill("solid", fgColor=_ZEBRA_FILL)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autosize_columns(ws, df: pd.DataFrame) -> None:
    """Ajusta largura de cada coluna com base no conteudo (max 42, min 13)."""
    for idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).head(200)])
        ws.column_dimensions[col_letter].width = min(max(13, max_len + 3), 42)


def _apply_number_formats(
    ws,
    columns: list,
    date_cols: Iterable[str] = (),
    int_cols: Iterable[str] = (),
    float_cols: Iterable[str] = (),
) -> None:
    """Aplica formatos de numero/data a colunas especificas da planilha."""
    header_map = {name: idx + 1 for idx, name in enumerate(columns)}
    formats = [(date_cols, "DD/MM/YYYY"), (int_cols, "#,##0"), (float_cols, "#,##0.00")]
    for cols, number_format in formats:
        for col_name in cols:
            col_idx = header_map.get(col_name)
            if not col_idx:
                continue
            col_letter = get_column_letter(col_idx)
            for row in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row}"].number_format = number_format


# --------------------------------------------------------------------------- #
# API publica
# --------------------------------------------------------------------------- #
def build_grouped_summary(
    df: pd.DataFrame,
    date_col: str,
    oficina_col: str,
    sum_cols: Iterable[str],
) -> pd.DataFrame:
    """
    Agrupa o DataFrame por Data+Oficina, somando as colunas numericas indicadas.
    Usado como aba principal do Excel: gestores consomem por dia/oficina.
    """
    sum_cols = list(sum_cols)
    grouped = (
        df.copy()
        .groupby([date_col, oficina_col], as_index=False)[sum_cols]
        .sum()
        .sort_values([date_col, oficina_col])
        .reset_index(drop=True)
    )
    return grouped


def to_excel_bytes_grouped(
    detail_df: pd.DataFrame,
    oficina_col: str,
    date_col: Optional[str] = None,
    int_cols: Optional[Iterable[str]] = None,
    float_cols: Optional[Iterable[str]] = None,
    summary_sheet: str = "Resumo Diario",
    detail_sheet: str = "Detalhado",
) -> bytes:
    """
    Gera um .xlsx em memoria com uma ou duas abas:
    - Resumo Diario (so quando date_col for fornecido): agrupado por Data+Oficina,
      linha TOTAL com formula SUM.
    - Detalhado: registro a registro igual ao exibido na tela.
    Retorna bytes prontos para st.download_button.
    """
    int_cols = list(int_cols or [])
    float_cols = list(float_cols or [])
    sum_cols = int_cols + float_cols

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        if date_col is not None:
            summary_df = build_grouped_summary(detail_df, date_col, oficina_col, sum_cols)

            summary_df.to_excel(writer, index=False, sheet_name=summary_sheet)
            ws_sum = writer.sheets[summary_sheet]
            _apply_number_formats(
                ws_sum, list(summary_df.columns),
                date_cols=[date_col], int_cols=int_cols, float_cols=float_cols,
            )

            total_row_idx = len(summary_df) + 2
            header_map = {name: idx + 1 for idx, name in enumerate(summary_df.columns)}
            ws_sum.cell(row=total_row_idx, column=1, value="TOTAL")
            for col_name in sum_cols:
                col_idx = header_map.get(col_name)
                if not col_idx:
                    continue
                col_letter = get_column_letter(col_idx)
                ws_sum.cell(
                    row=total_row_idx, column=col_idx,
                    value=f"=SUM({col_letter}2:{col_letter}{total_row_idx - 1})",
                )
                ws_sum[f"{col_letter}{total_row_idx}"].number_format = (
                    "#,##0" if col_name in int_cols else "#,##0.00"
                )

            _autosize_columns(ws_sum, summary_df)
            _style_worksheet(
                ws_sum,
                n_data_rows=total_row_idx - 1,
                n_cols=len(summary_df.columns),
                total_row=total_row_idx,
            )

        # aba detalhada
        detail_df.to_excel(writer, index=False, sheet_name=detail_sheet)
        ws_det = writer.sheets[detail_sheet]
        _apply_number_formats(
            ws_det, list(detail_df.columns),
            date_cols=[date_col] if date_col else [], int_cols=int_cols, float_cols=float_cols,
        )
        _autosize_columns(ws_det, detail_df)
        _style_worksheet(ws_det, n_data_rows=len(detail_df), n_cols=len(detail_df.columns))

    buffer.seek(0)
    return buffer.getvalue()


def download_button(
    df: pd.DataFrame,
    filename: str,
    oficina_col: str,
    date_col: Optional[str] = None,
    int_cols: Optional[Iterable[str]] = None,
    float_cols: Optional[Iterable[str]] = None,
    label: str = "Exportar Excel (resumo por dia/oficina + detalhado)",
    key: Optional[str] = None,
) -> None:
    """Renderiza o botao de download do Excel agrupado + detalhado."""
    if df.empty:
        return
    data = to_excel_bytes_grouped(
        df,
        oficina_col=oficina_col,
        date_col=date_col,
        int_cols=int_cols,
        float_cols=float_cols,
    )
    st.download_button(
        label,
        data=data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )
